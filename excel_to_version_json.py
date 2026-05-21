#!/usr/bin/env python3
"""Convert version.xlsx to version.json.

Expected sheets:
- 在线版本
- 审核版本
- 历史版本
Each sheet layout:
- row1: Chinese headers (ignored)
- row2: English keys: version, coverageApp, date, remark, user
- row3+: data
"""

import argparse
import json
from pathlib import Path
from openpyxl import load_workbook

FIELDS = ["version", "coverageApp", "date", "remark", "user"]
HEADER_ROW = 2
DATA_START_ROW = 3


def normalize(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def empty_str_if_none(v):
    return "" if v is None else v


def read_rows(ws):
    headers = [normalize(c.value) for c in ws[HEADER_ROW]]
    headers = [h for h in headers if h is not None]
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        vals = [normalize(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        if not any(v is not None for v in vals):
            continue
        row = dict(zip(headers, vals))
        item = {k: empty_str_if_none(row.get(k)) for k in FIELDS}
        rows.append(item)
    return rows


def first_or_empty(lst):
    return lst[0] if lst else {}


def main():
    parser = argparse.ArgumentParser(description="Convert version.xlsx to version.json")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    wb = load_workbook(in_path, data_only=True)

    def get_sheet(name):
        return wb[name] if name in wb.sheetnames else None

    online_rows = read_rows(get_sheet("在线版本")) if get_sheet("在线版本") else []
    audit_rows = read_rows(get_sheet("审核版本")) if get_sheet("审核版本") else []
    history_rows = read_rows(get_sheet("历史版本")) if get_sheet("历史版本") else []

    result = {
        "online": first_or_empty(online_rows),
        "audit": first_or_empty(audit_rows),
        "history": history_rows,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=4) + "\n", encoding="utf-8")

    print(f"written: {out_path}")
    print(f"online rows: {len(online_rows)}")
    print(f"audit rows: {len(audit_rows)}")
    print(f"history rows: {len(history_rows)}")


if __name__ == "__main__":
    main()
