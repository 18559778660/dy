#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook

RETENTION_FIELDS = [
    'date','appId','appName','os','factor',
    'activeUsers','newUsers','activePaidUsers','newPaidUsers',
    'day1','day2','day3','day4','day5','day6','day7','day14','day30'
]
SIDEBAR_FIELDS = [
    'date','appId','appName','os',
    'activeUsers','newUsers','activePaidUsers','newPaidUsers','penetrationRate',
    'day1','day2','day3','day4','day5','day6','day7','day14','day30'
]
HEADER_ROW=2
DATA_START=3


def norm(v):
    if v is None:
        return None
    if isinstance(v,str):
        s=v.strip()
        return s if s else ''
    return v


def read_sheet(ws, fields):
    out=[]
    for r in range(DATA_START, ws.max_row+1):
        vals=[norm(ws.cell(r,c).value) for c in range(1,len(fields)+1)]
        if not any(v not in (None,'') for v in vals):
            continue
        item={k:v for k,v in zip(fields,vals)}
        out.append(item)
    return out


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--input',required=True)
    ap.add_argument('--output',required=True)
    ap.add_argument('--source-json',required=False)
    args=ap.parse_args()

    wb=load_workbook(args.input,data_only=True)
    ws1=wb['留存分析'] if '留存分析' in wb.sheetnames else None
    ws2=wb['侧边栏留存'] if '侧边栏留存' in wb.sheetnames else None

    retention = read_sheet(ws1, RETENTION_FIELDS) if ws1 else []
    sidebar = read_sheet(ws2, SIDEBAR_FIELDS) if ws2 else []

    desc = {}
    if args.source_json:
        p=Path(args.source_json)
        if p.is_file():
            try:
                old=json.loads(p.read_text(encoding='utf-8'))
                desc=old.get('_说明') or {}
            except Exception:
                desc={}

    result={
        '_说明': desc,
        'retentionData': retention,
        'sidebarRetentionData': sidebar,
    }

    out=Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result,ensure_ascii=False,indent=2)+"\n",encoding='utf-8')
    print(f'written: {out}')
    print(f'retention rows: {len(retention)}')
    print(f'sidebar rows: {len(sidebar)}')


if __name__=='__main__':
    main()
