#!/usr/bin/env python3
"""Convert chart-data.xlsx (行为分析/实时分析) into chart-data.json structure.

Usage:
  python3 excel_to_chart_data.py \
    --input /path/chart-data.xlsx \
    --output /path/chart-data.json
"""

import argparse
import json
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


BEHAVIOR_SHEET = "行为分析"
REALTIME_SHEET = "实时分析"
HEADER_ROW = 2   # row 1 is Chinese labels, row 2 is English field keys
DATA_START_ROW = 3
MAIN_METRIC_FIELDS = [
    "dailyUsers",
    "newUsers",
    "avgDuration",
    "sharing",
    "startup",
    "avgStartup",
    "singleAvgDuration",
    "shareSuccess",
    "shareNewUsers",
    "shareSuccessUsers",
    "totalShares",
]


def normalize(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def to_int_if_possible(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            f = float(s)
            if f.is_integer():
                return int(f)
            return f
        except ValueError:
            return s
    return v


def read_sheet_as_dict_rows(ws):
    headers = [normalize(c.value) for c in ws[HEADER_ROW]]
    headers = [h for h in headers if h is not None]
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        vals = [normalize(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


def build_overview(behavior_rows):
    # app_key -> list[data_item]
    apps = OrderedDict()
    current_ctx = None  # { app_key, date, os }

    for row in behavior_rows:
        # Ignore completely empty rows
        if not any(normalize(v) is not None for v in row.values()):
            continue

        raw_app_id = normalize(row.get("appId"))
        raw_app_name = normalize(row.get("appName"))

        # Anchor row: has appId => always starts a new independent data item
        if raw_app_id is not None:
            app_id = raw_app_id
            app_name = raw_app_name or app_id
            app_key = (app_id, app_name)
            if app_key not in apps:
                apps[app_key] = []

            day_item = {
                "date": normalize(row.get("date")),
                "os": normalize(row.get("os")),
                "dailyUsers": to_int_if_possible(row.get("dailyUsers")),
                "newUsers": to_int_if_possible(row.get("newUsers")),
                "avgDuration": to_int_if_possible(row.get("avgDuration")),
                "sharing": to_int_if_possible(row.get("sharing")),
                "startup": to_int_if_possible(row.get("startup")),
                "avgStartup": to_int_if_possible(row.get("avgStartup")),
                "singleAvgDuration": to_int_if_possible(row.get("singleAvgDuration")),
                "shareSuccess": to_int_if_possible(row.get("shareSuccess")),
                "shareNewUsers": to_int_if_possible(row.get("shareNewUsers")),
                "shareSuccessUsers": to_int_if_possible(row.get("shareSuccessUsers")),
                "totalShares": to_int_if_possible(row.get("totalShares")),
                "douyinSourceScenes": [],
                "douyinVideoData": [],
            }
            apps[app_key].append(day_item)
            current_ctx = {
                "app_key": app_key,
                "date": day_item.get("date"),
                "os": day_item.get("os"),
            }
        else:
            # Child row: no appId but has any field => create a NEW data item
            # under the previous anchor app group.
            if not current_ctx:
                continue
            app_key = current_ctx["app_key"]
            day_item = {
                "date": normalize(row.get("date")),
                "os": normalize(row.get("os")),
                "dailyUsers": to_int_if_possible(row.get("dailyUsers")),
                "newUsers": to_int_if_possible(row.get("newUsers")),
                "avgDuration": to_int_if_possible(row.get("avgDuration")),
                "sharing": to_int_if_possible(row.get("sharing")),
                "startup": to_int_if_possible(row.get("startup")),
                "avgStartup": to_int_if_possible(row.get("avgStartup")),
                "singleAvgDuration": to_int_if_possible(row.get("singleAvgDuration")),
                "shareSuccess": to_int_if_possible(row.get("shareSuccess")),
                "shareNewUsers": to_int_if_possible(row.get("shareNewUsers")),
                "shareSuccessUsers": to_int_if_possible(row.get("shareSuccessUsers")),
                "totalShares": to_int_if_possible(row.get("totalShares")),
                "douyinSourceScenes": [],
                "douyinVideoData": [],
            }
            apps[app_key].append(day_item)
            current_ctx = {
                "app_key": app_key,
                "date": day_item.get("date"),
                "os": day_item.get("os"),
            }

        # source scene (optional per row)
        scene_id = normalize(row.get("SourceScenes_sceneId"))
        scene_name = normalize(row.get("SourceScenes_sceneName"))
        if scene_id is not None or scene_name is not None:
            scene_item = {
                "sceneId": scene_id,
                "sceneName": scene_name,
                "dailyUsers": to_int_if_possible(row.get("SourceScenes_dailyUsers")),
                "newUsers": to_int_if_possible(row.get("SourceScenes_newUsers")),
                "startup": to_int_if_possible(row.get("SourceScenes_startup")),
                "singleAvgDuration": to_int_if_possible(row.get("SourceScenes_singleAvgDuration")),
            }
            if scene_item not in day_item["douyinSourceScenes"]:
                day_item["douyinSourceScenes"].append(scene_item)

        # video data (optional per row)
        v_type = normalize(row.get("VideoData_type"))
        if v_type is not None:
            video_item = {
                "type": v_type,
                "newUsers": to_int_if_possible(row.get("VideoData_newUsers")),
                "activeUsers": to_int_if_possible(row.get("VideoData_activeUsers")),
                "conversionRate": to_int_if_possible(row.get("VideoData_conversionRate")),
            }
            if video_item not in day_item["douyinVideoData"]:
                day_item["douyinVideoData"].append(video_item)

    overview = []
    for (app_id, app_name), data_list in apps.items():
        overview.append({
            "appId": app_id,
            "appName": app_name,
            "data": data_list,
        })

    return overview


def build_realtime(realtime_rows):
    out = []
    for row in realtime_rows:
        app_id = normalize(row.get("appId"))
        app_name = normalize(row.get("appName"))
        if not app_id:
            continue
        out.append({
            "appId": app_id,
            "appName": app_name or app_id,
            "dailyVisitors": to_int_if_possible(row.get("dailyVisitors")),
            "dailyVisits": to_int_if_possible(row.get("dailyVisits")),
        })
    return out


def main():
    parser = argparse.ArgumentParser(description="Convert chart-data.xlsx to chart-data.json")
    parser.add_argument("--input", required=True, help="Input xlsx path")
    parser.add_argument("--output", required=True, help="Output json path")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    wb = load_workbook(input_path, data_only=True)

    if BEHAVIOR_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {BEHAVIOR_SHEET}")
    if REALTIME_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {REALTIME_SHEET}")

    behavior_rows = read_sheet_as_dict_rows(wb[BEHAVIOR_SHEET])
    realtime_rows = read_sheet_as_dict_rows(wb[REALTIME_SHEET])

    result = {
        "overview": build_overview(behavior_rows),
        "realTime": build_realtime(realtime_rows),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"written: {output_path}")
    print(f"overview apps: {len(result['overview'])}")
    print(f"realTime rows: {len(result['realTime'])}")


if __name__ == "__main__":
    main()
